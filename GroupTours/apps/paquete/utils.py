"""
Utilidades para manejo de conversión de monedas en paquetes.
"""
from decimal import Decimal, InvalidOperation
from io import BytesIO
from django.core.exceptions import ValidationError
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def _to_decimal(value):
    """
    Convierte un valor a Decimal de forma segura.

    Args:
        value: Valor a convertir (puede ser string, int, float, Decimal, None)

    Returns:
        Decimal: Valor convertido, o Decimal("0") si no es válido
    """
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value).replace(",", "."))
    except (InvalidOperation, ValueError, TypeError):
        return Decimal("0")


def convertir_entre_monedas(monto, moneda_origen, moneda_destino, fecha=None):
    """
    Convierte un monto entre USD y PYG en ambas direcciones.

    Esta función maneja las conversiones bidireccionales entre dólares y guaraníes
    utilizando las cotizaciones vigentes para una fecha específica.

    Args:
        monto: Decimal/float/int - Monto a convertir
        moneda_origen: Moneda - Moneda del monto original
        moneda_destino: Moneda - Moneda a la cual convertir
        fecha: date (opcional) - Fecha para obtener cotización. Si es None, usa fecha actual.

    Returns:
        Decimal: Monto convertido a la moneda destino

    Raises:
        ValidationError: Si no existe cotización vigente para la fecha especificada
        ValidationError: Si se intenta convertir entre monedas no soportadas

    Examples:
        >>> usd = Moneda.objects.get(codigo='USD')
        >>> pyg = Moneda.objects.get(codigo='PYG')
        >>> convertir_entre_monedas(100, usd, pyg)  # 100 USD → PYG
        Decimal('730000.00')
        >>> convertir_entre_monedas(730000, pyg, usd)  # 730000 PYG → USD
        Decimal('100.00')
    """
    from apps.moneda.models import Moneda, CotizacionMoneda

    # Si las monedas son iguales, retornar el mismo monto
    if moneda_origen == moneda_destino:
        return _to_decimal(monto)

    # Usar fecha actual si no se especifica
    if fecha is None:
        fecha = timezone.now().date()

    monto_decimal = _to_decimal(monto)

    # ========== CASO 1: Convertir de USD a PYG ==========
    if moneda_origen.codigo == 'USD' and moneda_destino.codigo == 'PYG':
        cotizacion = CotizacionMoneda.obtener_cotizacion_vigente(moneda_origen, fecha)

        if not cotizacion:
            raise ValidationError(
                f"No existe cotización de USD vigente para {fecha.strftime('%d/%m/%Y')}. "
                "Por favor registre una cotización antes de continuar."
            )

        valor_cotizacion = _to_decimal(cotizacion.valor_en_guaranies)
        return monto_decimal * valor_cotizacion

    # ========== CASO 2: Convertir de PYG a USD ==========
    if moneda_origen.codigo == 'PYG' and moneda_destino.codigo == 'USD':
        # Necesitamos la cotización de USD (cuántos guaraníes vale 1 dólar)
        moneda_usd = Moneda.objects.get(codigo='USD')
        cotizacion = CotizacionMoneda.obtener_cotizacion_vigente(moneda_usd, fecha)

        if not cotizacion:
            raise ValidationError(
                f"No existe cotización de USD vigente para {fecha.strftime('%d/%m/%Y')}. "
                "Por favor registre una cotización antes de continuar."
            )

        valor_cotizacion = _to_decimal(cotizacion.valor_en_guaranies)
        return monto_decimal / valor_cotizacion

    # ========== Cualquier otra combinación no está soportada ==========
    raise ValidationError(
        f"Conversión no soportada entre {moneda_origen.codigo} y {moneda_destino.codigo}. "
        "Solo se admiten conversiones entre USD y PYG."
    )


def generar_excel_pasajeros_salida(salida, pasajeros):
    """
    Genera un Excel con el listado de pasajeros de una salida.

    Args:
        salida: instancia de SalidaPaquete
        pasajeros: lista de dicts con datos de pasajeros (misma estructura que el endpoint)

    Returns:
        BytesIO con el archivo Excel
    """
    buffer = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Pasajeros"

    # ----- Encabezado de la salida -----
    ws["A1"] = f"Pasajeros - {salida.paquete.nombre}"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:N1")

    ws["A2"] = f"Salida: {salida.codigo}  |  Fecha: {salida.fecha_salida}"
    ws["A2"].font = Font(italic=True)
    ws.merge_cells("A2:N2")

    # ----- Headers -----
    headers = [
        "N°", "Reserva", "Estado Reserva", "Titular",
        "Nombre", "Apellido", "Tipo Doc.", "Documento",
        "Fecha Nac.", "Edad", "Precio Asignado", "Monto Pagado",
        "Saldo Pendiente", "Voucher",
    ]
    ws.append([])  # fila 3 vacía
    ws.append(headers)  # fila 4

    header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ----- Datos -----
    for idx, p in enumerate(pasajeros, 1):
        ws.append([
            idx,
            p.get("reserva_codigo", ""),
            p.get("reserva_estado_display") or p.get("reserva_estado", ""),
            "Sí" if p.get("es_titular") else "No",
            "Por asignar" if p.get("por_asignar") else (p.get("nombre", "") or ""),
            "Por asignar" if p.get("por_asignar") else (p.get("apellido", "") or ""),
            p.get("tipo_documento", "") or "",
            "Por asignar" if p.get("por_asignar") else (p.get("documento", "") or ""),
            p["fecha_nacimiento"].strftime("%d/%m/%Y") if p.get("fecha_nacimiento") else "",
            p.get("edad", "") if p.get("edad") is not None else "",
            float(p["precio_asignado"]) if p.get("precio_asignado") is not None else "",
            float(p["monto_pagado"]) if p.get("monto_pagado") is not None else "",
            float(p["saldo_pendiente"]) if p.get("saldo_pendiente") is not None else "",
            p.get("voucher_codigo", "") or "",
        ])

    # Formato montos
    for row in ws.iter_rows(min_row=5, min_col=11, max_col=13):
        for cell in row:
            if cell.value != "":
                cell.number_format = "#,##0.00"

    # Auto-ajustar ancho de columnas (omitir MergedCell que no tienen column_letter)
    # Se omiten las filas 1-3 (títulos fusionados) para que no inflen el ancho de columna A
    from openpyxl.cell.cell import MergedCell
    for col in ws.columns:
        anchor = next((cell for cell in col if not isinstance(cell, MergedCell)), None)
        if anchor is None:
            continue
        max_len = max(
            (len(str(cell.value or "")) for cell in col if not isinstance(cell, MergedCell) and cell.row >= 4),
            default=0,
        )
        ws.column_dimensions[anchor.column_letter].width = min(max_len + 3, 40)

    ws.auto_filter.ref = f"A4:{ws.cell(row=4, column=len(headers)).coordinate}"

    wb.save(buffer)
    buffer.seek(0)
    return buffer


def obtener_cotizacion_actual(moneda):
    """
    Obtiene la cotización vigente actual para una moneda.

    Args:
        moneda: Moneda - Moneda a consultar (debe ser USD)

    Returns:
        CotizacionMoneda o None si no existe cotización vigente
    """
    from apps.moneda.models import CotizacionMoneda

    if moneda.codigo == 'PYG':
        return None  # Guaraníes no tiene cotización (es la moneda base)

    return CotizacionMoneda.obtener_cotizacion_vigente(moneda)
