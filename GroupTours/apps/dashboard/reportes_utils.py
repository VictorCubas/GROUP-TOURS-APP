"""
Utilidades para exportación de reportes a PDF y Excel.
"""
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.enums import TA_CENTER, TA_RIGHT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================================
# EXPORTACIÓN PDF
# ============================================================================

def generar_pdf_movimientos_cajas(data, filtros, resumen):
    """
    Genera PDF del reporte de movimientos de cajas.
    
    Args:
        data: Lista de movimientos (ya serializados)
        filtros: Dict con filtros aplicados
        resumen: Dict con totales
    
    Returns:
        BytesIO con el PDF
    """
    buffer = BytesIO()
    
    # Crear documento (horizontal)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Contenido
    story = []
    
    # Título
    story.append(Paragraph("REPORTE DE MOVIMIENTOS DE CAJAS", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Información de filtros
    filtros_text = f"<b>Período:</b> {filtros['fecha_desde']} a {filtros['fecha_hasta']}"
    if filtros.get('caja_id'):
        filtros_text += f" | <b>Caja:</b> {filtros['caja_id']}"
    story.append(Paragraph(filtros_text, styles['Normal']))
    story.append(Spacer(1, 0.3*cm))
    
    # Resumen
    resumen_text = f"""
    <b>Total Ingresos:</b> ₲ {format_money(resumen['total_ingresos'])} 
    ({format_usd(resumen.get('total_ingresos_usd'))} USD) | 
    <b>Total Egresos:</b> ₲ {format_money(resumen['total_egresos'])} 
    ({format_usd(resumen.get('total_egresos_usd'))} USD) | 
    <b>Balance:</b> ₲ {format_money(resumen['balance'])} 
    ({format_usd(resumen.get('balance_usd'))} USD)
    """
    story.append(Paragraph(resumen_text, styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Tabla de datos
    table_data = [
        ['N°', 'Fecha/Hora', 'Caja', 'Tipo', 'Concepto', 'Monto ₲', 'Monto USD', 'Método', 'Usuario']
    ]
    
    for idx, mov in enumerate(data, 1):
        monto_usd = mov.get('monto_usd', 0) or 0
        table_data.append([
            str(idx),
            format_datetime(mov['fecha_hora']),
            mov['caja_nombre'],
            mov['tipo_movimiento_display'],
            mov['concepto_display'][:25],  # Truncar
            f"₲ {format_money(mov['monto'])}",
            f"$ {format_usd(monto_usd)}",
            mov['metodo_pago_display'][:8],
            mov['usuario_registro'][:15]  # Truncar
        ])
    
    # Crear tabla
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (5, 1), (5, -1), 'RIGHT'),  # Monto alineado derecha
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Zebra stripes
        *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ecf0f1')) 
          for i in range(2, len(table_data), 2)]
    ]))
    
    story.append(table)
    
    # Footer
    story.append(Spacer(1, 1*cm))
    footer_text = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_pdf_paquetes(data, filtros, resumen):
    """Genera PDF del reporte de paquetes."""
    buffer = BytesIO()
    
    # Crear documento (horizontal)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Contenido
    story = []
    
    # Título
    story.append(Paragraph("REPORTE DE PAQUETES TURÍSTICOS", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Información de filtros
    filtros_text = f"<b>Estado:</b> {filtros.get('estado', 'todos')}"
    if filtros.get('fecha_desde'):
        filtros_text += f" | <b>Desde:</b> {filtros['fecha_desde']}"
    if filtros.get('fecha_hasta'):
        filtros_text += f" | <b>Hasta:</b> {filtros['fecha_hasta']}"
    story.append(Paragraph(filtros_text, styles['Normal']))
    story.append(Spacer(1, 0.3*cm))
    
    # Resumen
    resumen_text = f"""
    <b>Total Paquetes:</b> {resumen['total_registros']} | 
    <b>Activos:</b> {resumen['paquetes_activos']} | 
    <b>Precio Promedio (PYG):</b> {format_money(resumen.get('precio_promedio_pyg') or 0)} | 
    <b>Precio Promedio (USD):</b> $ {format_usd(resumen.get('precio_promedio_usd') or 0)}
    """
    story.append(Paragraph(resumen_text, styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Tabla de datos
    table_data = [
        ['Código', 'Nombre', 'Destino', 'Precio ₲', 'Precio USD', 'Cupos', 'Reservas', 'Estado']
    ]
    
    for paq in data:
        precio_gs = paq.get('precio_gs', 0) or 0
        precio_usd = paq.get('precio_usd', 0) or 0
        # Si el paquete NO es propio, mostrar "N/A" en cupos (sujeto a disponibilidad)
        cupos_text = 'N/A' if not paq.get('propio') else str(paq.get('cupos_disponibles', 0) or 0)
        
        table_data.append([
            paq['codigo'],
            paq['nombre'][:30],  # Truncar
            paq['destino_ciudad'],
            f"₲ {format_money(precio_gs)}",
            f"$ {format_usd(precio_usd)}",
            cupos_text,
            str(paq['reservas_count']),
            'Activo' if paq['activo'] else 'Inactivo'
        ])
    
    # Crear tabla
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Precio
        ('ALIGN', (4, 1), (6, -1), 'CENTER'),  # Cupos, Reservas, Estado
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Zebra stripes
        *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ecf0f1')) 
          for i in range(2, len(table_data), 2)]
    ]))
    
    story.append(table)
    
    # Footer
    story.append(Spacer(1, 1*cm))
    footer_text = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_pdf_reservas(data, filtros, resumen):
    """Genera PDF del reporte de reservas."""
    buffer = BytesIO()
    
    # Crear documento (horizontal)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1*cm,
        leftMargin=1*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    # Contenido
    story = []
    
    # Título
    story.append(Paragraph("REPORTE DE RESERVAS", title_style))
    story.append(Spacer(1, 0.5*cm))
    
    # Información de filtros
    filtros_text = f"<b>Estado:</b> {filtros.get('estado', 'todas')}"
    if filtros.get('fecha_desde'):
        filtros_text += f" | <b>Desde:</b> {filtros['fecha_desde']}"
    if filtros.get('fecha_hasta'):
        filtros_text += f" | <b>Hasta:</b> {filtros['fecha_hasta']}"
    story.append(Paragraph(filtros_text, styles['Normal']))
    story.append(Spacer(1, 0.3*cm))
    
    # Resumen
    resumen_text = f"""
    <b>Total Reservas:</b> {resumen['total_registros']} | 
    <b>Monto Total (₲):</b> {format_money(resumen['monto_total'])} | 
    <b>Monto Total (USD):</b> $ {format_usd(resumen.get('monto_total_usd'))} | 
    <b>Saldo Pendiente (₲):</b> {format_money(resumen['saldo_pendiente'])} | 
    <b>Saldo Pendiente (USD):</b> $ {format_usd(resumen.get('saldo_pendiente_usd'))}
    """
    story.append(Paragraph(resumen_text, styles['Normal']))
    story.append(Spacer(1, 0.5*cm))
    
    # Tabla de datos
    table_data = [
        ['Código', 'Titular', 'Paquete', 'F. Salida', 'Pax', 'Monto ₲', 'Monto USD', 'Pagado %', 'Estado']
    ]
    
    for res in data:
        monto_usd = res.get('monto_total_usd', 0) or 0
        table_data.append([
            res['codigo'],
            res['titular_nombre'][:20],  # Truncar
            res['paquete_nombre'][:25],  # Truncar
            res['fecha_salida'][:10] if res.get('fecha_salida') else 'N/A',
            str(res['cantidad_pasajeros']),
            f"₲ {format_money(res['monto_total'])}",
            f"$ {format_usd(monto_usd)}",
            f"{res['porcentaje_pagado']:.0f}%",
            res['estado_display']
        ])
    
    # Crear tabla
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Body
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('ALIGN', (4, 1), (6, -1), 'CENTER'),  # Pax, Monto, Pagado
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        
        # Zebra stripes
        *[('BACKGROUND', (0, i), (-1, i), colors.HexColor('#ecf0f1')) 
          for i in range(2, len(table_data), 2)]
    ]))
    
    story.append(table)
    
    # Footer
    story.append(Spacer(1, 1*cm))
    footer_text = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    story.append(Paragraph(footer_text, styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    buffer.seek(0)
    return buffer


# ============================================================================
# EXPORTACIÓN EXCEL
# ============================================================================

def generar_excel_movimientos_cajas(data, filtros, resumen):
    """
    Genera Excel del reporte de movimientos de cajas.
    
    Returns:
        BytesIO con el archivo Excel
    """
    buffer = BytesIO()
    wb = Workbook()
    
    # ===== HOJA 1: RESUMEN =====
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Título
    ws_resumen['A1'] = "REPORTE DE MOVIMIENTOS DE CAJAS"
    ws_resumen['A1'].font = Font(size=16, bold=True)
    ws_resumen.merge_cells('A1:D1')
    
    # Filtros
    ws_resumen['A3'] = "Período:"
    ws_resumen['B3'] = f"{filtros['fecha_desde']} a {filtros['fecha_hasta']}"
    
    # Resumen
    ws_resumen['A5'] = "Total Ingresos (PYG):"
    ws_resumen['B5'] = float(resumen['total_ingresos'])
    ws_resumen['B5'].number_format = '#,##0.00'
    
    ws_resumen['A6'] = "Total Ingresos (USD):"
    ws_resumen['B6'] = float(resumen.get('total_ingresos_usd') or 0)
    ws_resumen['B6'].number_format = '#,##0.00'
    
    ws_resumen['A7'] = "Total Egresos (PYG):"
    ws_resumen['B7'] = float(resumen['total_egresos'])
    ws_resumen['B7'].number_format = '#,##0.00'
    
    ws_resumen['A8'] = "Total Egresos (USD):"
    ws_resumen['B8'] = float(resumen.get('total_egresos_usd') or 0)
    ws_resumen['B8'].number_format = '#,##0.00'
    
    ws_resumen['A9'] = "Balance (PYG):"
    ws_resumen['B9'] = float(resumen['balance'])
    ws_resumen['B9'].number_format = '#,##0.00'
    ws_resumen['B9'].font = Font(bold=True)
    
    ws_resumen['A10'] = "Balance (USD):"
    ws_resumen['B10'] = float(resumen.get('balance_usd') or 0)
    ws_resumen['B10'].number_format = '#,##0.00'
    ws_resumen['B10'].font = Font(bold=True)
    
    # ===== HOJA 2: DATOS =====
    ws_datos = wb.create_sheet("Datos")
    
    # Headers
    headers = ['N°', 'Fecha/Hora', 'Caja', 'Tipo', 'Concepto', 'Descripción', 
               'Monto (₲)', 'Monto (USD)', 'Método Pago', 'Referencia', 'Usuario']
    ws_datos.append(headers)
    
    # Estilo header
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_datos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for idx, mov in enumerate(data, 1):
        monto_usd = mov.get('monto_usd', 0) or 0
        ws_datos.append([
            idx,
            format_datetime(mov['fecha_hora']),
            mov['caja_nombre'],
            mov['tipo_movimiento_display'],
            mov['concepto_display'],
            mov.get('descripcion', '') or '',
            float(mov['monto']),
            float(monto_usd),
            mov['metodo_pago_display'],
            mov.get('referencia', '') or '',
            mov['usuario_registro']
        ])
    
    # Formato columnas de montos
    for row in range(2, len(data) + 2):
        ws_datos[f'G{row}'].number_format = '#,##0.00'  # Monto PYG
        ws_datos[f'H{row}'].number_format = '#,##0.00'  # Monto USD
    
    # Auto-ajustar columnas
    for column in ws_datos.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_datos.column_dimensions[column_letter].width = adjusted_width
    
    # Filtros en Excel
    ws_datos.auto_filter.ref = ws_datos.dimensions
    
    # Guardar
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_paquetes(data, filtros, resumen):
    """Genera Excel del reporte de paquetes."""
    buffer = BytesIO()
    wb = Workbook()
    
    # ===== HOJA 1: RESUMEN =====
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Título
    ws_resumen['A1'] = "REPORTE DE PAQUETES TURÍSTICOS"
    ws_resumen['A1'].font = Font(size=16, bold=True)
    ws_resumen.merge_cells('A1:D1')
    
    # Resumen
    ws_resumen['A3'] = "Total Paquetes:"
    ws_resumen['B3'] = resumen['total_registros']
    
    ws_resumen['A4'] = "Paquetes Activos:"
    ws_resumen['B4'] = resumen['paquetes_activos']
    
    ws_resumen['A5'] = "Paquetes Inactivos:"
    ws_resumen['B5'] = resumen['paquetes_inactivos']
    
    ws_resumen['A7'] = "Precio Promedio (PYG):"
    ws_resumen['B7'] = float(resumen.get('precio_promedio_pyg') or 0)
    ws_resumen['B7'].number_format = '#,##0.00'
    
    ws_resumen['A8'] = "Precio Promedio (USD):"
    ws_resumen['B8'] = float(resumen.get('precio_promedio_usd') or 0)
    ws_resumen['B8'].number_format = '#,##0.00'
    
    # ===== HOJA 2: DATOS =====
    ws_datos = wb.create_sheet("Datos")
    
    # Headers
    headers = ['Código', 'Nombre', 'Tipo', 'Destino', 'País', 'Distribuidora', 
               'Precio (₲)', 'Precio (USD)', 'Seña (₲)', 'Seña (USD)', 'Moneda Orig.', 
               'Fecha Inicio', 'Duración (días)', 'Cupos Disponibles', 'Cupos Ocupados', 
               'Reservas', 'Personalizado', 'Propio', 'Estado']
    ws_datos.append(headers)
    
    # Estilo header
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_datos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for paq in data:
        precio_gs = paq.get('precio_gs', 0) or 0
        precio_usd = paq.get('precio_usd', 0) or 0
        sena_gs = paq.get('sena_gs', 0) or 0
        sena_usd = paq.get('sena_usd', 0) or 0
        
        # Si el paquete NO es propio, mostrar "N/A" en cupos (sujeto a disponibilidad)
        cupos_disp = 'N/A' if not paq.get('propio') else (paq.get('cupos_disponibles', 0) or 0)
        cupos_ocup = 'N/A' if not paq.get('propio') else (paq.get('cupos_ocupados', 0) or 0)
        
        ws_datos.append([
            paq['codigo'],
            paq['nombre'],
            paq['tipo_paquete'],
            paq['destino_ciudad'],
            paq['destino_pais'],
            paq.get('distribuidora', 'N/A') or 'N/A',
            float(precio_gs),
            float(precio_usd),
            float(sena_gs),
            float(sena_usd),
            paq['moneda'],
            paq.get('fecha_inicio', ''),
            paq.get('duracion_dias', 0),
            cupos_disp,
            cupos_ocup,
            paq['reservas_count'],
            'Sí' if paq['personalizado'] else 'No',
            'Sí' if paq['propio'] else 'No',
            'Activo' if paq['activo'] else 'Inactivo'
        ])
    
    # Formato columnas
    for row in range(2, len(data) + 2):
        ws_datos[f'G{row}'].number_format = '#,##0.00'  # Precio PYG
        ws_datos[f'H{row}'].number_format = '#,##0.00'  # Precio USD
        ws_datos[f'I{row}'].number_format = '#,##0.00'  # Seña PYG
        ws_datos[f'J{row}'].number_format = '#,##0.00'  # Seña USD
    
    # Auto-ajustar columnas
    for column in ws_datos.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_datos.column_dimensions[column_letter].width = adjusted_width
    
    # Filtros en Excel
    ws_datos.auto_filter.ref = ws_datos.dimensions
    
    # Guardar
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generar_excel_reservas(data, filtros, resumen):
    """Genera Excel del reporte de reservas."""
    buffer = BytesIO()
    wb = Workbook()
    
    # ===== HOJA 1: RESUMEN =====
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    
    # Título
    ws_resumen['A1'] = "REPORTE DE RESERVAS"
    ws_resumen['A1'].font = Font(size=16, bold=True)
    ws_resumen.merge_cells('A1:D1')
    
    # Resumen
    ws_resumen['A3'] = "Total Reservas:"
    ws_resumen['B3'] = resumen['total_registros']
    
    ws_resumen['A4'] = "Reservas Pendientes:"
    ws_resumen['B4'] = resumen['reservas_pendientes']
    
    ws_resumen['A5'] = "Reservas Confirmadas:"
    ws_resumen['B5'] = resumen['reservas_confirmadas']
    
    ws_resumen['A6'] = "Reservas Finalizadas:"
    ws_resumen['B6'] = resumen['reservas_finalizadas']
    
    ws_resumen['A7'] = "Reservas Canceladas:"
    ws_resumen['B7'] = resumen['reservas_canceladas']
    
    ws_resumen['A9'] = "Monto Total (PYG):"
    ws_resumen['B9'] = float(resumen['monto_total'])
    ws_resumen['B9'].number_format = '#,##0.00'
    
    ws_resumen['A10'] = "Monto Total (USD):"
    ws_resumen['B10'] = float(resumen.get('monto_total_usd') or 0)
    ws_resumen['B10'].number_format = '#,##0.00'
    
    ws_resumen['A11'] = "Monto Pagado (PYG):"
    ws_resumen['B11'] = float(resumen['monto_pagado'])
    ws_resumen['B11'].number_format = '#,##0.00'
    
    ws_resumen['A12'] = "Monto Pagado (USD):"
    ws_resumen['B12'] = float(resumen.get('monto_pagado_usd') or 0)
    ws_resumen['B12'].number_format = '#,##0.00'
    
    ws_resumen['A13'] = "Saldo Pendiente (PYG):"
    ws_resumen['B13'] = float(resumen['saldo_pendiente'])
    ws_resumen['B13'].number_format = '#,##0.00'
    ws_resumen['B13'].font = Font(bold=True)
    
    ws_resumen['A14'] = "Saldo Pendiente (USD):"
    ws_resumen['B14'] = float(resumen.get('saldo_pendiente_usd') or 0)
    ws_resumen['B14'].number_format = '#,##0.00'
    ws_resumen['B14'].font = Font(bold=True)
    
    # ===== HOJA 2: DATOS =====
    ws_datos = wb.create_sheet("Datos")
    
    # Headers
    headers = ['Código', 'Fecha Reserva', 'Titular', 'Documento', 'Email', 'Teléfono', 'Paquete', 
               'Destino', 'Fecha Salida', 'Fecha Retorno', 'Pasajeros', 
               'Precio Unit. (₲)', 'Precio Unit. (USD)', 
               'Monto Total (₲)', 'Monto Total (USD)', 
               'Monto Pagado (₲)', 'Monto Pagado (USD)', 
               'Saldo Pend. (₲)', 'Saldo Pend. (USD)', 
               '% Pagado', 'Estado', 'Estado Pago', 'Modalidad Fact.']
    ws_datos.append(headers)
    
    # Estilo header
    header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws_datos[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for res in data:
        precio_usd = res.get('precio_unitario_usd', 0) or 0
        monto_total_usd = res.get('monto_total_usd', 0) or 0
        monto_pagado_usd = res.get('monto_pagado_usd', 0) or 0
        saldo_usd = res.get('saldo_pendiente_usd', 0) or 0
        
        ws_datos.append([
            res['codigo'],
            res['fecha_reserva'][:10] if res.get('fecha_reserva') else '',
            res['titular_nombre'],
            res.get('titular_documento', ''),
            res.get('titular_email', ''),
            res.get('titular_telefono', ''),
            res['paquete_nombre'],
            res['destino_completo'],
            res.get('fecha_salida', ''),
            res.get('fecha_retorno', ''),
            res['cantidad_pasajeros'],
            float(res['precio_unitario']),
            float(precio_usd),
            float(res['monto_total']),
            float(monto_total_usd),
            float(res['monto_pagado']),
            float(monto_pagado_usd),
            float(res['saldo_pendiente']),
            float(saldo_usd),
            float(res['porcentaje_pagado']),
            res['estado_display'],
            res['estado_pago_display'],
            res['modalidad_facturacion_display']
        ])
    
    # Formato columnas
    for row in range(2, len(data) + 2):
        ws_datos[f'L{row}'].number_format = '#,##0.00'  # Precio Unit PYG
        ws_datos[f'M{row}'].number_format = '#,##0.00'  # Precio Unit USD
        ws_datos[f'N{row}'].number_format = '#,##0.00'  # Monto Total PYG
        ws_datos[f'O{row}'].number_format = '#,##0.00'  # Monto Total USD
        ws_datos[f'P{row}'].number_format = '#,##0.00'  # Monto Pagado PYG
        ws_datos[f'Q{row}'].number_format = '#,##0.00'  # Monto Pagado USD
        ws_datos[f'R{row}'].number_format = '#,##0.00'  # Saldo Pend PYG
        ws_datos[f'S{row}'].number_format = '#,##0.00'  # Saldo Pend USD
        ws_datos[f'T{row}'].number_format = '0.00'      # % Pagado
    
    # Auto-ajustar columnas
    for column in ws_datos.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_datos.column_dimensions[column_letter].width = adjusted_width
    
    # Filtros en Excel
    ws_datos.auto_filter.ref = ws_datos.dimensions
    
    # Guardar
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================================
# UTILIDADES
# ============================================================================

def format_money(value):
    """Formatea un número como dinero."""
    try:
        num = float(value)
        return f"{num:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except:
        return "0,00"


def format_usd(value):
    """Formatea un número como dólares."""
    try:
        if value is None:
            return "N/A"
        num = float(value)
        return f"{num:,.2f}"
    except:
        return "N/A"


def format_datetime(value):
    """Formatea un datetime."""
    try:
        if isinstance(value, str):
            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
        else:
            dt = value
        return dt.strftime('%d/%m/%Y %H:%M')
    except:
        return str(value)

